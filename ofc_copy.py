from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Style, fills, PatternFill, Color
import datetime, time
import os, xlrd

coldic = {'Cocoa':'D2691E','Aqua':'7EC0EE','Olive':'89892B','Yellow':'CCA300'} 

def compare(s1, s2, w1n,w2n, col):
    
    rowrange = max(s1.max_row, s2.max_row)
    colrange = max(s1.max_column, s2.max_column)
    cfill = Style(fill=PatternFill(patternType='solid', fgColor=Color(coldic[col])))
    succ = "Pass"
    e = 'Data Matches!'
    try:
        for i in xrange(1,rowrange+1):
            for j in xrange(1,colrange+1):
                #time.sleep(0.2)
                a = s1.cell(row=i,column=j).value
                b = s2.cell(row=i,column=j).value
                nt = None
                dt = datetime.datetime.today()
                tlist = [type(int()), type(float()),type(long()), type(nt)]
                numlist = [type(int()), type(float()),type(long())]
                for_code = s2.cell(row=i,column=j).number_format
                if type(a) not in numlist and type(b) not in numlist:
                    a = unicode(a)
                    b = unicode(b)
                if a != b:
                    succ = "Fail"
                    e = "Data Mismatches"
                    comtxt = None
                    
                    if type(a) in numlist and (type(b) == type(nt) or type(b) == type(unicode())):
                        comtxt = str(w1n)+": " + unicode(a)+ ". diff: "+ unicode(a)
                    
                    elif type(b) in numlist and (type(a) == type(nt) or type(a) == type(unicode())):
                        comtxt = str(w1n)+": " + unicode(a)+ ". diff: "+ unicode(b)
                    
                    elif type(a) in numlist and type(b) in numlist:
                        comtxt = str(w1n)+": " + unicode(a)+ ". diff: "+ unicode(b-a)

                    elif type(a) == type(unicode()) and type(b) == type(unicode()):
                        comtxt = str(w1n)+": " + unicode(a)+ "."
                
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).style = cfill
                    s2.cell(row=i,column=j).comment = comment
                    s2.cell(row=i,column=j).number_format = for_code
                else:
                    comtxt = None
    except KeyboardInterrupt:
        succ = "Abort"
        e = "KeyboardInterrupt"
    except Exception as e:
        succ = "Abort"
            
    return [succ,e]
              

def the_mess(l):
    col = l[3]
    print l
    report = []
    if l[1] != '' and l[2] != '':
        f1 = l[1]
        f2 = l[2]
        work_dir = l[2][:l[2].rfind('/')].replace('/','//')
        os.chdir(work_dir)
        rprt = the_mayhem(f1,f2,col)
        report.append(rprt)
    elif l[0] != '':
        work_dir = l[0].replace("/","//")
        os.chdir(work_dir)
        print "successfully changed path!"
        files_list = filter(lambda x: x.endswith('.xlsx') or x.endswith('.xls'), os.listdir(work_dir))
        print files_list
        lfile = len(files_list)
        cnt = 0
        while cnt < lfile:
            f1 = files_list[cnt]
            f2 = files_list[cnt+1]
            cnt += 2
            rprt = the_mayhem(f1,f2,col)
            report.append(rprt)
    return report
    #print report

def the_mayhem(f1,f2,col):
    print f1, f2
    t1 = time.time()
    if f1.endswith('.xls') and f2.endswith('.xls'):
        w1 = open_xls_as_xlsx(f1)
        w2 = open_xls_as_xlsx(f2)
    else:
        w1 = load_workbook(f1)
        w2 = load_workbook(f2)
    w1n = f1[f1.rfind('/')+1:f1.find('.')]
    w2n = f2[f2.rfind('/')+1:f2.find('.')]
    wbsucc = "Pass"
    sdic = {}

    for i in xrange(len(w1.worksheets)):
        for j in xrange(i,i+1):
            smsg = compare(w1.worksheets[j], w2.worksheets[j],w1n,w2n,col)
            sdic[w1.worksheets[j].title] = smsg[0],smsg[1]
            if f2.endswith('.xls'):
                w2.save(w2n + str('.xlsx'))
            else:
                w2.save(f2)
    for e in sdic.values():
        if "Pass" not in e[0]:
            wbsucc = "Fail"
    
    t2 = time.time()
    ttime = round((t2-t1),2)
    #print sdic
    return [ttime,w1n,w2n,wbsucc,sdic]
   
def open_xls_as_xlsx(filename):
    # first open using xlrd
    ob = xlrd.open_workbook(filename, formatting_info=True)
    nb = Workbook()
    nb.remove_sheet(nb.worksheets[0])
    index = 0
    nsheets = ob.nsheets
    
    while index < ob.nsheets:
        os = ob.sheet_by_index(index)
        nrows = os.nrows
        ncols = os.ncols

        # prepare a xlsx sheet
        ns = nb.create_sheet(index)
        ns.title = os.name
        #print ns.title,os.name
        for row in xrange(1, nrows+1):
            for col in xrange(1, ncols+1):
                nc = ns.cell(row=row, column=col)
                oc = os.cell(row-1, col-1)
                nc.value = oc.value
                old_format = ob.format_map[ob.xf_list[oc.xf_index].format_key].format_str
                nc.number_format = old_format
        index += 1    
        #nb.save("c:\\users\\nikil\\Desktop\\sample1234.xlsx")
    return nb


if __name__ == '__main__':
    main()
