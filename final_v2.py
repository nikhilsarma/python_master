"""
__author__ = "Nikhil" [Suman (Idea) Kushal, Sushant]
__copyright__ = "LoL"
__version__ = "1.0"
__maintainer__ = "Team E"
__status__ = "Mess"

Hits the Database, Executes the queries and formulates a report based on the outputs
can execute simultaneously across two environments of Databases
"""


import pyodbc
import time,datetime
import os,xlsxwriter,time
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Style, fills, PatternFill, Color
from exp_final_v2 import *



"""
Function to take care of teh Singleton result sets
A singleton resultSet --> output is 1X1 (mostly a count check)
Returns the Testcase Number, Test Condition, Execution status, Message(Pass/Fail),
Counts across 2 Environments / Exception message if any,
Time taken to execute the query

Report Format: Same as the I/P file with Extra columns for:
Execution status,
Message(Pass/Fail),
Counts across 2 Environments / Exception message if any,
Time taken to execute the query

"""

def count_check(tc_no,tc,c_qry,cur1,cur2):
	
	try:
                t1 = time.time()
                c1,c2 = cur1.execute(c_qry),cur2.execute(c_qry)
                a,b = c1.fetchone(),c2.fetchone()
                a1 = 0 if type(a).__name__ == 'NoneType' else a.time.strftime("%d-%b-%Y") if type(a).__name__ == 'date' else a[0]
                b1 = 0 if type(b).__name__ == 'NoneType' else a.time.strftime("%d-%b-%Y") if type(a).__name__ == 'date' else b[0]
                #print a1,b1
		t2 = time.time()
		tt = round((t2-t1),2)
	except Exception as e:
		msg = "Fail"
		ex_stat = "N"
		
		return [tc_no,tc,ex_stat,msg,(e.__class__.__name__,e.args),str(round((time.time()-t1),2))]
	else:
                
                msg = "Pass" if a == b else "Fail"			
		ex_stat = "Y"
		#print "hi nikhil"
		#print [tc_no,tc,ex_stat,msg,(str(a1),str(b1)),str(tt)]
		return [tc_no,tc,ex_stat,msg,(str(a1),str(b1)),str(tt)]



"""
Function to take care of Exportd & Data compared result sets
An Exported & Data compared resultSet --> Oupput of Two Excel sheets from diff. environments - Data compared 
Returns the Testcase Number, Test Condition, Execution status, Message(Pass/Fail),
Counts across 2 Environments / Exception message if any,
Time taken to execute the query

Report Format: Same as the I/P file with Extra columns for:
Execution status,
Message(Pass/Fail) - With a link to open the Datacompared file on the Disk
Counts across 2 Environments / Exception message if any,
Time taken to execute the query

"""

def re_runner(base1,base2,regfile,totes):
        
        db1,db2 = base1,base2
        cnxn_1,cnxn_2 = pyodbc.connect('DSN='+db1), pyodbc.connect('DSN='+db2)
        cur1,cur2= cnxn_1.cursor(),cnxn_2.cursor()
        print "connection Established"
        frr = regfile
        wd= frr[:frr.rfind('/')].replace('/','//')
        os.chdir(wd)

        wbrr = load_workbook(frr)
        wkrr = xlsxwriter.Workbook('dbtest_rerun_check.xlsx', {'constant_memory': True})
        bold_f = wkrr.add_format({'bold': True})#1e321e
        pass_format = wkrr.add_format({'font_color': 'black', 'indent': '1','bg_color': '99CC00', 'italic': 1, 'underline': 1,'font_size':  12,})
        fail_format = wkrr.add_format({'font_color': 'black', 'indent': '1','bg_color': 'FF7C80', 'italic': 1, 'underline': 1,'font_size':  12,})
        red_format = wkrr.add_format({'font_color': 'black', 'indent': '1','bg_color': 'FF7C80', 'italic': 1,})
        green_format = wkrr.add_format({'font_color': 'black','indent': '1', 'bg_color': '99CC00', 'italic': 1,})
        total_result_rr = []
        #wsht = wkrr.add_worksheet('Re_Run')
        xl_hd = ["Test_Case_No","Test_Condition","SQL Query","Executed","Result","O/P: Comments","Time Taken(sec..)"]
        rtxt = "Test Case   Exe.Stat   Result   O/P : Comments"
        qq = len(rtxt)*"-"
        print "\n" + "Pulihooara!"
        print  rtxt + "\n" + qq
        tobef = totes
        toaft = []
        for e in wbrr.worksheets:
                ridx = 0
                for i in range(1,e.max_row+1):
                        if e.cell(row=i,column=1).value in tobef:
                                tobef.remove(e.cell(row=i,column=1).value),toaft.append(e.cell(row=i,column=1).value)
                                ridx += 1
                                if e.title not in wkrr.sheetnames:
                                        wsht = wkrr.add_worksheet(e.title)
                                        for hdr in enumerate(xl_hd):
                                                wsht.write(0,hdr[0],hdr[1],bold_f)
                                else:
                                        wsht.sheet_name = e.title
                                #print "manipulating the sheet :" + e.title + " at the row :" + str(i)
                                tcaseno = e.cell(row=i,column=1).value
                                tcond = e.cell(row=i,column=2).value
                                squery = e.cell(row=i,column=3).value
                                if tcond[tcond.rfind('_')+1:] == '1':
                                        res = count_check(tcaseno,tcond,squery.replace('\n',' '),cur1,cur2)
                                        res.insert(2,squery)
                                        total_result_rr.append(res)
                                        #print res
                                        for col in enumerate(res):
                                                if col[0] == 5:
                                                        if res[3] == "Y":
                                                                cmt = db1+": " + str(col[1][0]) + " ; " + db2+": " + str(col[1][1])
                                                        else:
                                                                cmt = str(col[1][0]) +" : " + str(col[1][1][1].replace(' ',''))
                                                wsht.write(ridx,col[0],cmt if col[0] == 5 else col[1])
                                                if col[0] == 3:
                                                        if col[1] == "Y":
                                                                wsht.write(ridx,3, "Y", green_format)
                                                        else:
                                                                wsht.write(ridx,3, "N", red_format)

                                                if col[0] == 4:
                                                        if col[1] == "Pass":
                                                                wsht.write(ridx,4, "Pass", green_format)
                                                        else:
                                                                wsht.write(ridx,4, "Fail", red_format)
                                                                        
                                        print res[0] + res[-4].rjust(10) + res[-3].rjust(13) + "\t" + cmt
                                        
                                else:
                                        res = all_check(tcaseno,tcond,squery.replace('\n',' '),cur1,cur2,db1,db2)
                                        rows = res[1]
                                        es = "N" if (rows[0][0] == -1 or rows[1][0] == -1) else "Y"
                                        flink = res[0][1]
                                        linkfile = res[0][1]
                                        tip = linkfile[linkfile.rfind("/")+1:]
                                        result = res[0][0][0][3] if es == "Y" else "Fail"
                                        v_comment = res[0][0][0][4] if result == "Verify" else ''
                                        fnl_re = [tcaseno,tcond,squery.replace('\n',' '),es,result,'',str(res[2])]
                                        #print fnl_re
                                        total_result_rr.append(fnl_re)
                                        for col in enumerate(fnl_re):
                                                if col[0] == 5:
                                                        if es == "Y":
                                                                if result == "Verify":
                                                                        cmt = db1+": " + str(rows[0][0]) + " ; " + db2+": " + str(rows[1][0]) + "\t " + str(v_comment)
                                                                else:
                                                                        cmt = db1+": " + str(rows[0][0]) + " ; " + db2+": " + str(rows[1][0])
                                                        else:
                                                                cmt = str(rows[0][1][0]) + ": " + str(rows[0][1][1][1].replace(' ',''))
                                                wsht.write(ridx,col[0],cmt if col[0] == 5 else col[1])
                                                if col[0] == 3:
                                                        if es == "Y":
                                                                wsht.write(ridx,3, "Y", green_format)
                                                        else:
                                                                wsht.write(ridx,3, "N", red_format)

                                                if col[0] == 4:
                                                        if result == "Pass":
                                                                wsht.write_url(ridx,4, linkfile, pass_format, result, tip)
                                                        else:
                                                                wsht.write_url(ridx,4, linkfile, fail_format, result, tip)
                                                
                                        print fnl_re[0] + fnl_re[-4].rjust(10) + fnl_re[-3].rjust(13) + "\t" + cmt.strip('\n')




def runto_excel(base1,base2,regfile):
    
    """
    what? :
    Takes in the Regresion file,
    Creates the connections Cnxn_1/2 and Associated cursors cur1/2 (Pointers to the Output resultset)
    Changes the working Directory(wd) to the current file location
    Opens a Workbook and writes the result to it.
    Parallely spits out the results onto the Shell

    Logic :
    Checks for the final value on the Test condition
    _1 - COunt check query and hence a singleton result set --> gives to the "count_check()"
    __ - Everything Else (mostly needs an extract from Db and hence given to the "export_check"
    
    Finally: The result is stored in "res" and is iterated through to print to the Excel
    Openpyxl writes to the Excel in a row=0,col=0 basis (Hence the Enemurate function in for loop - to get the row-index)
    Using two loop conditions
    [Repeat for the result]: One row --> multiple columns
    Check for the column type and write with according formats
    Except for Exceptions if any and close the workbook
    """      
    
    db1,db2 = base1,base2
    #print db1,db2
    cnxn_1,cnxn_2 = pyodbc.connect('DSN='+db1), pyodbc.connect('DSN='+db2)
    #print "hehehehe"
    cur1,cur2= cnxn_1.cursor(),cnxn_2.cursor()
    print "connection Established"

    f = regfile
    wd= f[:f.rfind('/')].replace('/','//')
    os.chdir(wd)
    try:
            #opens a workbook and setsup the formats as necessary
            wb1 = load_workbook(f,use_iterators = True)
            workbook = xlsxwriter.Workbook('dbtest_count_check.xlsx', {'constant_memory': True})
            bold_f = workbook.add_format({'bold': True})#1e321e
            pass_format = workbook.add_format({'font_color': 'black', 'indent': '1','bg_color': '99CC00', 'italic': 1, 'underline': 1,'font_size':  12,})
            fail_format = workbook.add_format({'font_color': 'black', 'indent': '1','bg_color': 'FF7C80', 'italic': 1, 'underline': 1,'font_size':  12,})
            red_format = workbook.add_format({'font_color': 'black', 'indent': '1','bg_color': 'FF7C80', 'italic': 1,})
            green_format = workbook.add_format({'font_color': 'black','indent': '1', 'bg_color': '99CC00', 'italic': 1,})
            total_result = []
            tone = time.time()
            #The column names for the O/P workbook and Shell 
            xl_hd = ["Test_Case_No","Test_Condition","SQL Query","Executed","Result","O/P: Comments","Time Taken(sec..)"]
            rtxt = "Test Case   Exe.Stat   Result   O/P : Comments"
            under_line = len(rtxt)*"-"
            for e in wb1.worksheets:
                ws = e
                row = ws.iter_rows()
                row.next()
                s_name = ws.title
                worksheet = workbook.add_worksheet(s_name)

                print "\n" + s_name
                print  rtxt + "\n" + under_line
                for e in enumerate(xl_hd):
                    worksheet.write(0,e[0],e[1],bold_f)
                for e in enumerate(row):
                    if e[1][1].value[e[1][1].value.rfind('_')+1:] == "1":
                        #res = count_check(e[0].value,e[1].value,e[2].value.replace('\n',''))
                        res = count_check(e[1][0].value,e[1][1].value,e[1][2].value.replace('\n',' '),cur1,cur2)
                        res.insert(2,e[1][2].value)
                        #print res
                        total_result.append(res)

                        for col in enumerate(res):
                            #print col
                            if col[0] == 5:
                                    if res[3] == "Y":
                                            cmt = db1+": " + str(col[1][0]) + " ; " + db2+": " + str(col[1][1])
                                    else:
                                            cmt = str(col[1][0]) +" : " + str(col[1][1][1].replace(' ',''))
                            worksheet.write(e[0]+1,col[0],cmt if col[0] == 5 else col[1])
                            if col[0] == 3:
                                    if col[1] == "Y":
                                            worksheet.write(e[0]+1,3, "Y", green_format)
                                    else:
                                            worksheet.write(e[0]+1,3, "N", red_format)

                            if col[0] == 4:
                                    if col[1] == "Pass":
                                            worksheet.write(e[0]+1,4, "Pass", green_format)
                                    else:
                                            worksheet.write(e[0]+1,4, "Fail", red_format)

                        print res[0] + res[-4].rjust(10) + res[-3].rjust(13) + "\t" + cmt
                    else:
                            #print "now in else about to go to all_check function"
                          
                            res = export_check(e[1][0].value,e[1][1].value,e[1][2].value.replace('\n',' '),cur1,cur2,db1,db2)
                            print res
                            rows = res[1]
                            es = "N" if (rows[0][0] == -1 or rows[1][0] == -1) else "Y"
                            flink = res[0][1]
                            linkfile = flink[flink.rfind('Extracts'):]
                            print linkfile
                            tip = linkfile[linkfile.rfind("/")+1:]
                            result = res[0][0][0][3] if es == "Y" else "Fail"
                            v_comment = res[0][0][0][4] if result == "Verify" else ''
                            fnl_re = [e[1][0].value,e[1][1].value,e[1][2].value.replace('\n',' '),es,result,'',str(res[2])]
                            #print fnl_re
                            total_result.append(fnl_re)

                            for col in enumerate(fnl_re):
                                    if col[0] == 5:
                                            if es == "Y":
                                                    if result == "Verify":
                                                            cmt = db1+": " + str(rows[0][0]) + " ; " + db2+": " + str(rows[1][0]) + "\t " + str(v_comment)
                                                    else:
                                                            cmt = db1+": " + str(rows[0][0]) + " ; " + db2+": " + str(rows[1][0])
                                            else:
                                                    if rows[0][0] == -1:
                                                            cmt = str(rows[0][1][0]) + ": " + str(rows[0][1][1][1].replace(' ',''))
                                                    elif rows[1][0] == -1:
                                                            cmt = str(rows[1][1][0]) + ": " + str(rows[1][1][1][1].replace(' ',''))
                                    worksheet.write(e[0]+1,col[0],cmt if col[0] == 5 else col[1])
                                    if col[0] == 3:
                                            if es == "Y":
                                                    worksheet.write(e[0]+1,3, "Y", green_format)
                                            else:
                                                    worksheet.write(e[0]+1,3, "N", red_format)

                                    if col[0] == 4:
                                            if result == "Pass":
                                                    worksheet.write_url(e[0]+1,4, linkfile, pass_format, result, tip)
                                            else:
                                                    worksheet.write_url(e[0]+1,4, linkfile, fail_format, result, tip)
                                                  
                            print fnl_re[0] + fnl_re[-4].rjust(10) + fnl_re[-3].rjust(13) + "\t" + cmt.strip('\n')
                            
    except KeyboardInterrupt:
        print "program stopped by user"
        workbook.close()
        ttwo = time.time()
        print "total time is " + str(round((ttwo-tone),2)) + "sec"
        return total_result

    except Exception as e3:
            workbook.close()
            print "program Stopped Abruptly : " + str(e3.__class__.__name__) + str(e)
            ttwo = time.time()
            print "\ntotal time is " + str(round((ttwo-tone),2)) + "sec"
            return total_result

    workbook.close()
    ttwo = time.time()
    print "total time is " + str(round((ttwo-tone),2)) + "sec"
    return total_result                
                                     
               
