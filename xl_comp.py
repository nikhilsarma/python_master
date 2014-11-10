from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Style, fills, PatternFill, Color
import datetime, time
import os

def compare(s1, s2, w1n,w2n):
    print "hello nikhil"
    pass
    rowrange = max(s1.max_row, s2.max_row)
    colrange = max(s1.max_column, s2.max_column)
    cfill = Style(fill=PatternFill(patternType='solid', fgColor=Color('00ff00')))
    
    for i in range(1,rowrange+1):
        for j in range(1,colrange+1):
            a = s1.cell(row=i,column=j).value
            b = s2.cell(row=i,column=j).value
            nt = None
            dt = datetime.datetime.today()
            tlist = [type(int()), type(float()),type(long()), type(nt)]
            numlist = [type(int()), type(float()),type(long())]
                
            if type(a) not in numlist and type(b) not in numlist:
                a = str(unicode(a))
                b = str(unicode(b))
            if a != b:
                s2.cell(row=i,column=j).style = cfill
                if type(a) in numlist and type(b) == type(nt):
#                    print "hi" + str(a)
                    comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(a)
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(b) in numlist and type(a) == type(nt):
#                    print "hi" + str(b)
                    comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(b)
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(a) in numlist and type(b) in numlist:
                    comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(b-a)
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(a) == type(nt) and type(b) == type(str()):
                    comtxt = str(w1n)+": " + str(a)+ "."
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(b) == type(nt) and type(a) == type(str()):
                    comtxt = str(w1n)+": " + str(a)+ "."
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(a) == type(str()) and type(b) == type(str()):
                    comtxt = str(w1n)+": " + str(a)+ "."
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
"""
                elif type(a) == type(dt) and type(b) == type(dt):
                    comtxt = str(w1n)+": " + str(a.date())+ "."
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
"""

#sets the path to the curret working directory where the files to be compared can be found easily
#path = raw_input("enter the path of present working directory: ")                

def the_mess(l):
    print l
    if l[0] != '':
        work_dir = l[0].replace("/","//")      
    else:
        work_dir = l[1][:l[1].rfind('/')].replace('/','//')
        print work_dir
    if os.getcwd() != work_dir:
        os.chdir(work_dir)
        print "successfully changed path!"
    if l[1] != '' and l[2] != '':
        f1 = l[1][l[1].rfind('/')+1:]
        f2 = l[2][l[2].rfind('/')+1:]
        the_mayhem(f1,f2)
    elif l[0] != '':
        files_list = filter(lambda x: x.endswith('.xlsx'), os.listdir(work_dir))
        print files_list
        lfile = len(files_list)
        cnt = 0
        while cnt < lfile:
            f1 = files_list[cnt]
            f2 = files_list[cnt+1]
            cnt += 2
            the_mayhem(f1,f2)

def the_mayhem(f1,f2):
    print f1, f2
    t1 = time.time()
    w1 = load_workbook(f1)
    w2 = load_workbook(f2)
    w1n = f1[:f1.find('.')]
    w2n = f2[:f2.find('.')]
#Gives the sheet number of the workbook by index
#s1 = w1.worksheets[0]
#s2 = w2.worksheets[0]
    for i in range(len(w1.worksheets)):
        for j in range(i,i+1):
#           print "comparing for sheet " + str(j+1)+ "..."
            compare(w1.worksheets[j], w2.worksheets[j],w1n,w2n)
            w2.save(f2)
#           print "the end"
    t2 = time.time()
    ttime = round((t2-t1),2)
    print "total time of execution is: " +str(ttime) + "sec."

if __name__ == '__main__':
    main()
