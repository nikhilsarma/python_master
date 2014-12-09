import datetime, time,traceback,xlsxwriter
from openpyxl import load_workbook
f1 = "c:/users/nikil/desktop/sap1.xlsx"
f2 = "c:/users/nikil/desktop/sap2.xlsx"
fname = "c:/users/nikil/desktop/sap123.xlsx"
t1 = time.time()
wb1 = load_workbook(f1,use_iterators = True)
wb2 = load_workbook(f2,use_iterators = True)
ws1 = wb1.worksheets[0]
ws2 = wb2.worksheets[0]
r1 = ws1.iter_rows()
r2 = ws2.iter_rows()
hrow = max(ws1.get_highest_row(),ws2.get_highest_row())
hcol = max(ws1.get_highest_column(),ws2.get_highest_column())
workbook = xlsxwriter.Workbook(fname, {'constant_memory': True})
worksheet = workbook.add_worksheet()
formt = workbook.add_format()
formt.set_bg_color('orange')
cnt = 1
while cnt <= ws1.get_highest_row():
    try:
        rone = r1.next()
        rtwo = r2.next()
        for i in xrange(ws1.get_highest_column()):
            a,b = rone[i].value, rtwo[i].value
            worksheet.write(cnt-1,i,b)
            if a!= b:
                #print rone[i].value
                #print "mismatch found at row: " +str(cnt)+" in cell: " +str(i+1)+"."
                #print str(a)+ " ---> " + str(b)
                worksheet.write(cnt-1,i,b,formt)
                worksheet.write_comment(cnt-1,i,str(a), {'x_scale': 0.7, 'y_scale': 0.6})
        cnt += 1
    
    except Exception,err:
        print traceback.format_exc()
        break
workbook.close()
t2 = time.time()
ttime = round((t2-t1),2)
print "total time of execution is: " +str(ttime) + "sec."

"""		
	except StopIteration:
		print "end of comp"
		t2 = time.time()
		ttime = round((t2-t1),2)
		print "total time of execution is: " +str(ttime) + "sec."
		break
"""
