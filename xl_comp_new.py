from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Style, fills, PatternFill, Color
import datetime, time
import os

def compare(s1, s2, w1n,w2n):
    #print "hello nikhil"
    #pass
    rowrange = max(s1.max_row, s2.max_row)
    colrange = max(s1.max_column, s2.max_column)
    cfill = Style(fill=PatternFill(patternType='solid', fgColor=Color('00ff00')))
    succ = "Pass"
    e = 'Data Matches!'
    try:
        for i in xrange(1,rowrange+1):
            for j in xrange(1,colrange+1):
                time.sleep(0.2)
                a = s1.cell(row=i,column=j).value
                b = s2.cell(row=i,column=j).value
                nt = None
                dt = datetime.datetime.today()
                tlist = [type(int()), type(float()),type(long()), type(nt)]
                numlist = [type(int()), type(float()),type(long())]
                
                if type(a) not in numlist and type(b) not in numlist:
                    a = str(a)
                    b = str(b)
                if a != b:
                    succ = "Fail"
                    e = "Data Mismatches"
                    #print a,b
                    #s2.cell(row=i,column=j).style = cfill
                    if type(a) in numlist and type(b) == type(nt):
                        #print "hi" + str(a)
                        comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(a)
                    
                    elif type(b) in numlist and type(a) == type(nt):
                        #print "hi" + str(b)
                        comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(b)
                    
                    elif type(a) in numlist and type(b) in numlist:
                        comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(b-a)

                    elif type(a) == type(str()) and type(b) == type(str()):
                        comtxt = str(w1n)+": " + str(a)+ "."

                    elif type(a) != type(b):
                        comtxt = str(w1n)+": " + str(a)+ ". \nIncompatible data"

                
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                    s2.cell(row=i,column=j).style = cfill
                else:
                    comtxt = None
    except KeyboardInterrupt:
        succ = "Abort"
        e = "KeyboardInterrupt"
        #print e
    except Exception as e:
        succ = "Abort"
        #e= e.message
        #print e
    
    return [succ,e]

#sets the path to the curret working directory where the files to be compared can be found easily
#path = raw_input("enter the path of present working directory: ")                

def the_mess(l):
    print l
    report = []
    if l[0] != '':
        work_dir = l[0].replace("/","//")      
    else:
        work_dir = l[1][:l[1].rfind('/')].replace('/','//')
        
    if os.getcwd() != work_dir:
        os.chdir(work_dir)
        print "successfully changed path!"
    if l[1] != '' and l[2] != '':
        f1 = l[1][l[1].rfind('/')+1:]
        f2 = l[2][l[2].rfind('/')+1:]
        rprt = the_mayhem(f1,f2)
        report.append(rprt)
    elif l[0] != '':
        files_list = filter(lambda x: x.endswith('.xlsx'), os.listdir(work_dir))
        #print files_list
        lfile = len(files_list)
        cnt = 0
        while cnt < lfile:
            f1 = files_list[cnt]
            f2 = files_list[cnt+1]
            cnt += 2
            rprt = the_mayhem(f1,f2)
            report.append(rprt)
    return report
    #print report

def the_mayhem(f1,f2):
    print f1, f2
    t1 = time.time()
    w1 = load_workbook(f1)
    w2 = load_workbook(f2)
    w1n = f1[:f1.find('.')]
    w2n = f2[:f2.find('.')]
    wbsucc = "Pass"
    sdic = {}
#Gives the sheet number of the workbook by index
#s1 = w1.worksheets[0]
#s2 = w2.worksheets[0]
    for i in xrange(len(w1.worksheets)):
        for j in xrange(i,i+1):
#           print "comparing for sheet " + str(j+1)+ "..."
            smsg = compare(w1.worksheets[j], w2.worksheets[j],w1n,w2n)
            sdic[w1.worksheets[j].title] = smsg[0],smsg[1]
            w2.save(f2)
    for e in sdic.values():
        if "Pass" not in e[0]:
            wbsucc = "Fail"
    
    t2 = time.time()
    ttime = round((t2-t1),2)
    #print sdic
    return [ttime,w1n,w2n,wbsucc,sdic]
    #print "total time of execution is: " +str(ttime) + "sec."

if __name__ == '__main__':
    main()
