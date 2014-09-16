
""" Python Script for comparing two excell sheets """

"""         Author: Nikhil Kumar Kadayinti

    1. Cmpared two .xlsx files cell to cell,
    2. Highlights the Mismatched Cell in the document with a fill color 
    3. Comments-out the difference among the compared cells.
        This script is built upon the Open-Source library "Openpyxl" used for manipulating Excell documents. Library "Jdcal" is required for calendar manipulations.
        
        A similar functionality can be achieved,using the libraries "Xlrd,Xwlt,Xutils" 

    Scope for improvisation:
    1. Reading data form Cells can be increased further with the help of iterators and generators, reducing the memory footprint.
    2. Can be brought in as a standalone application, UFriendly prompting to input the files for comparision
    3. If necessary teh sheets could be filtered accordingly( not possible with the present library, may be implemented in next patch for Openpyxl)
    3. Data comparision could get smart by including more data types and ignoring the differences, if reaching beyind a certain limit.(can use a
        small episilon variable to keep track of permissible difference limit. ex: e = 0.0001)
    4. handling None type objects are quite hard and hence making workaround for that. initialising nt to None and comparing with it all along.
    5. datetime objects are treated as strings and could be made visibly better.(though suffices)
"""    


from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Style, fills, PatternFill, Color
import datetime, time
import os

# Function to compare different Sheets of a Workbook
""" compares two sheets of different workbooks individually, diciding on which sheet has the maximum rows and which has the maximum columns, iterates over the
    entire permutable space.

    "cfill" variable used, along part of the styles module, creates a foreground color for the mismatched Cell in each sheet.
    "iterates over the sheet, visiting all the cells with the help of co-ordinates and takes the values to compare, casting all the values apart from INT and FLOAT
        to the type String for easy calculation of differences among the Numbers.
     Dealing with the type None is yet to be worked on. As of now None type is treated as "Zero".
     having Comments to each cell is achieved using the comments module of the library.
        Syntax: Comment("text", "author") creates the comment object, which can then be appended to the desired cell.
        
"""    
def compare(s1, s2, w1n,w2n):
    rowrange = max(s1.max_row, s2.max_row)
    colrange = max(s1.max_column, s2.max_column)
    cfill = Style(fill=PatternFill(patternType='solid', fgColor=Color('00ff00')))
#    print rowrange, colrange
    
    for i in range(1,rowrange+1):
        for j in range(1,colrange+1):
            a = s1.cell(row=i,column=j).value
            b = s2.cell(row=i,column=j).value
#            print a, type(a)
            nt = None
#            dt = datetime.datetime.today()
            tlist = [type(int()), type(float()),type(long()), type(nt)]
            numlist = [type(int()), type(float()),type(long())]
            if type(a) not in numlist and type(b) not in numlist:
                a = str(a)
                b = str(b)
#            print a,type(a)
#            print b,type(b)
            
            
            if a != b:
                
                if type(a) in numlist and type(b) == type(nt):
#                    print "hi" + str(a)
                    comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(a)
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(b) in numlist and type(a) == type(nt):
#                    print "hi" + str(b)
                    comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(b)
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(a) in numlist and type(b) in numlist:
                    comtxt = str(w1n)+": " + str(a)+ ". diff: "+ str(b-a)
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(a) == type(nt) and type(b) == type(str()):
                    comtxt = str(w1n)+": " + str(a)+ "."
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(b) == type(nt) and type(a) == type(str()):
                    comtxt = str(w1n)+": " + str(a)+ "."
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
                elif type(a) == type(str()) and type(b) == type(str()):
                    comtxt = str(w1n)+": " + str(a)+ "."
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
"""
                elif type(a) == type(dt) and type(b) == type(dt):
                    comtxt = str(w1n)+": " + str(a.date())+ "."
                    comment = Comment(comtxt, w2n)
                    s2.cell(row=i,column=j).comment = comment
"""
"""
#sets the path to the curret working directory where the files to be compared can be found easily
path = raw_input("enter the path of present working directory: ")                
if os.getcwd != path:
    os.chdir(path)
    print "successfully changed path!"
"""
#f1 = raw_input("enter the name of base file: ")                
#f2 = raw_input("enter the name of secondary file(updations are done to this): ")                

"""
Loads the respective Excell Workbook specified by the user and prepares the scene for comparing the workbook sheet by sheet, finally saving the workbook,
    with highlights for differences and comments duly put.

"""
f1 = 'sd_old.xlsx'
f2 = 'sd_new.xlsx'
t1 = time.time()
w1 = load_workbook(f1)
w2 = load_workbook(f2)
w1n = f1[:-5]
w2n = f2[:-5]
#Gives the sheet number of the workbook by index
#s1 = w1.worksheets[0]
#s2 = w2.worksheets[0]
for i in range(len(w1.worksheets)):
    for j in range(i,i+1):
#        print "comparing for sheet " + str(j+1)+ "..."
        compare(w1.worksheets[j], w2.worksheets[j],w1n,w2n)
        w2.save(f2)
        t2 = time.time()
#        print "the end"
ttime = round((t2-t1),2)
print "total time of execution is: " +str(ttime) + "sec."


