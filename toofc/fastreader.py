import datetime, time,traceback,xlsxwriter,os
from openpyxl import load_workbook
#f1 = "c:/users/nikil/desktop/ntest/vbf1.xlsx"
#f2 = "c:/users/nikil/desktop/ntest/vbf2.xlsx"
#fname = "c:/users/nikil/desktop/vbf123.xlsx"

def the_lmess(l):
    report = []
    if l[1] != '' and l[2] != '':
        f1,f2 = l[1],l[2]
        work_dir = l[2][:l[2].rfind('/')].replace('/','//')
        #print work_dir
        os.chdir(work_dir)
        rprt = comp_lxl(f1,f2)
        report.append(rprt)
    elif l[0] != '':
        work_dir = l[0].replace("/","//")
        os.chdir(work_dir)
        files_list = [e.lower() for e in filter(lambda x: x.endswith('.xlsx'), os.listdir(work_dir))]
        print files_list.sort()
        lfile = len(files_list)
        cnt = 0
        while cnt < lfile:
            f1 = files_list[cnt]
            f2 = files_list[cnt+1]
            cnt += 2
            rprt = comp_lxl(f1,f2)
            report.append(rprt)
    return report

def comp_lxl(f1,f2):
    sucmsg = "Passs"
    w1n = f1[f1.rfind('/')+1:f1.rfind('.')]
    w2n = f2[f2.rfind('/')+1:f2.rfind('.')]
    t1 = time.time()
    wb1 = load_workbook(f1,use_iterators = True)
    wb2 = load_workbook(f2,use_iterators = True)
    ws1 = wb1.worksheets[0]
    ws2 = wb2.worksheets[0]
    r1 = ws1.iter_rows()
    r2 = ws2.iter_rows()
    #hrow = max(ws1.get_highest_row(),ws2.get_highest_row())
    #hcol = max(ws1.get_highest_column(),ws2.get_highest_column())
    workbook = xlsxwriter.Workbook(w1n+str("_vs_")+w2n+str('.xlsx'), {'constant_memory': True})
    worksheet = workbook.add_worksheet()
    formt = workbook.add_format()
    formt.set_bg_color('orange')
    cnt = 1
    a = ws1.get_highest_row()
    print a
    #while cnt <= a:
    while cnt <= ws1.get_highest_row():
        try:
            rone,rtwo = r1.next(),r2.next()
            for j in xrange(ws1.get_highest_column()):
                a,b = rone[j].value, rtwo[j].value
                worksheet.write(cnt-1,j,b)
                if a != b:
                    sucmsg = "Fail"
                    #print rone[i].value
                    #print "mismatch found at row: " +str(cnt)+" in cell: " +str(i+1)+"."
                    #print str(a)+ " ---> " + str(b)
                    worksheet.write(cnt-1,j,b,formt)
                    worksheet.write_comment(cnt-1,j,str(a), {'x_scale': 0.7, 'y_scale': 0.6})
            cnt += 1
        
        except Exception,err:
            print traceback.format_exc()
            break
    workbook.close()
    t2 = time.time()
    td = round((t2-t1),2)
    return [td,w1n,w2n,sucmsg,w1n+str("_vs_")+w2n]
    print "total time of execution is: " +str(td) + "sec."

