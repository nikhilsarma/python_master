""" The mess1 and mayhem1: used for improvised xlx_xlsx conversion using the pywincom32 lib


"""

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Style, fills, PatternFill, Color
import datetime, time,os
import win32com.client
#import os, xlrd

coldic = {'Cocoa':'D2691E','Blue':'7EC0EE','Olive':'89892B','Yellow':'CCA300','Orange':'FF8000','Rose':'FF7C80','Green':'99CC00'} 

def compare(s1, s2, w1n,w2n):

    #rowrange = max(s1.get_highest_row(), s2.get_highest_row())
    #colrange = max(s1.get_highest_column(), s2.get_highest_column())
    #print rowrange,colrange
    #eps = 10
    rowrange = max(s1.max_row, s2.max_row)
    colrange = max(s1.max_column, s2.max_column)
    total_cells = rowrange*colrange
    fill1_err,fill2_err,fill3_err = 0,0,0
    diff_fill, unfill =0,total_cells
    #print rowrange,colrange,total_cells
    #print col1,col2,col3,val1,val2
    #pass

    cfill_1 = Style(fill=PatternFill(patternType='solid', fgColor=Color(coldic[col1])))
    cfill_2 = Style(fill=PatternFill(patternType='solid', fgColor=Color(coldic[col2])))
    cfill_3 = Style(fill=PatternFill(patternType='solid', fgColor=Color(coldic[col3])))
    cfill_4 = Style(fill=PatternFill(patternType='solid', fgColor=Color(coldic['Olive'])))
    succ = "Pass"
    e = 'Data Matches!'
    try:
        for i in xrange(1,rowrange+1):
            for j in xrange(1,colrange+1):
                #time.sleep(0.5)
                x = s1.cell(row=i,column=j)
                y = s2.cell(row=i,column=j)
                xfor_code,yfor_code = x.number_format,y.number_format
                
                a = x.value
                b = y.value
                nt = None
                dt = datetime.datetime.today()
                tlist = [type(int()), type(float()),type(long()), type(nt)]
                numlist = [type(int()), type(float()),type(long())]
                #xfor_code = s2.cell(row=i,column=j).number_format
                if (type(a) not in numlist or x.number_format == '@') and (type(b) not in numlist or y.number_format == '@'):
                    a,b = unicode(a).lower(),unicode(b).lower()
                    #print x.number_format+str('->')+y.number_format
                    #print a + str('->')+ b
                    #print str(type(a)) +str('->')+str(type(b))
                if a != b:
                    diff_fill += 1
                    #unfill -= 1
                    #print a,b
                    #print type(a),type(b)
                    succ = "Fail"
                    e = "Data Mismatches"
                    comtxt = None
                    
                    if type(a) in numlist and (type(b) == type(nt) or type(b) == type(unicode())):
                        comtxt = str(w1n)+": " + unicode(a)+ str("\ndiff: ")+ unicode(a)
                    
                    elif type(b) in numlist and (type(a) == type(nt) or type(a) == type(unicode())):
                        comtxt = str(w1n)+": " + unicode(a)+ str("\ndiff: ")+ unicode(b)
                    
                    elif type(a) in numlist and type(b) in numlist:
                        diff_val = float(abs(a-b))
                        diff_per = diff_val/abs(a)*100 if a!=0 else b
                        if diff_per <= float(eps):
                            continue
                        #y.style = cfill_1 if diff_per <= float(val1) else cfill_2 if diff_per <= float(val2) else cfill_3
                        if diff_per <= float(val1):
                            y.style = cfill_1
                            fill1_err += 1
                        elif diff_per <= float(val2):
                            y.style = cfill_2
                            fill2_err += 1
                        else:
                            y.style = cfill_3
                            fill3_err += 1

                        if a < b:
                            comtxt = str(w1n)+": " + unicode(a)+ str("\ndiff: ")+ unicode(b-a)+ str("\n")+unicode(round(diff_per,4))+ unicode(" % inc. " )
                        elif a > b:
                            comtxt = str(w1n)+": " + unicode(a)+ str("\ndiff: ")+ unicode(b-a)+str("\n")+unicode(round(diff_per,4))+ unicode(" % dec. " )
                        #comtxt = str(w1n)+": " + unicode(a)+ str("\ndiff: ")+ unicode(b-a)+ str("\ndiff %: ")+ unicode(diff_per)

                    elif type(a) == type(unicode()) and type(b) == type(unicode()):
                        comtxt = str(w1n)+": " + unicode(a)
                        y.style = cfill_4
                        fill3_err += colrange
                        #unfill -= colrange-1

                
                    comment = Comment(comtxt, w2n)
                    #y.style = cfill
                    y.comment = comment
                    y.number_format = yfor_code
                else:
                    comtxt = None
        print "nik"
        print fill1_err,fill2_err,fill3_err, total_cells,unfill,diff_fill
    except KeyboardInterrupt:
        succ = "Abort"
        e = "KeyboardInterrupt"
    except Exception as e:
        succ = "Abort"
            
    #print [succ,e]
    return [succ,e]
              

def the_mess(l):
    col,eps = l[3],l[4]
    print l
    print "Comparision started...!"
    report = []
    if l[1] != '' and l[2] != '':
        f1,f2 = l[1],l[2]
        work_dir = l[2][:l[2].rfind('/')].replace('/','//')
        #print work_dir
        os.chdir(work_dir)
        rprt = the_mayhem(f1,f2,col,eps)
        report.append(rprt)
    elif l[0] != '':
        work_dir = l[0].replace("/","//")
        os.chdir(work_dir)
        files_list = [e.lower() for e in filter(lambda x: x.endswith('.xlsx') or x.endswith('.xls'), os.listdir(work_dir))]
        files_list.sort()
        lfile = len(files_list)
        cnt = 0
        while cnt < lfile:
            f1 = files_list[cnt]
            f2 = files_list[cnt+1]
            cnt += 2
            rprt = the_mayhem(f1,f2,col,eps)
            report.append(rprt)
    return report
    #print report

def the_mayhem(f1,f2,col):
    print "Comparing... " +f1 +" vs " + f2
    t1 = time.time()
    if f1.endswith('.xls') and f2.endswith('.xls'):
        w1 = open_xls_as_xlsx(f1)
        w2 = open_xls_as_xlsx(f2)
    else:
        w1 = load_workbook(f1)
        w2 = load_workbook(f2)
    w1n = f1[f1.rfind('/')+1:f1.rfind('.')]
    w2n = f2[f2.rfind('/')+1:f2.rfind('.')]
    wbsucc = "Pass"
    sdic = {}

    for i in xrange(len(w1.worksheets)):
        for j in xrange(i,i+1):
            smsg = compare(w1.worksheets[j], w2.worksheets[j],w1n,w2n,col)
            sdic[w1.worksheets[j].title] = smsg[0],smsg[1]
            if f2.endswith('.xls') and smsg[0] == 'Fail':
                w2.save(w2n + str('.xlsx'))
            elif f2.endswith('.xlsx') and smsg[0] == 'Fail':
                w2.save(f2)
    for e in sdic.values():
        if "Pass" not in e[0]:
            wbsucc = "Fail"
    
    t2 = time.time()
    ttime = round((t2-t1),2)
    #print sdic
    return [ttime,w1n,w2n,wbsucc,sdic]




def xls_x_conv(xfl,*fpath):
    import shutil
    xl = win32com.client.Dispatch("Excel.Application")
    if fpath:
        os.chdir(fpath[0])
        print "in the conv function for.." + str(fpath[0])
        new_dir = os.getcwd()[os.getcwd().rfind("\\")+1:]
        if not os.path.exists(new_dir + "_xls_files"):
            os.makedirs(new_dir + "_xls_files")
        dst = new_dir + "_xls_files"
        for e in xfl:
            fname = os.path.join(os.getcwd(),e)
            wb = xl.Workbooks.Open(fname)
            wb.SaveAs(fname+"x", FileFormat = 51)
            wb.Close()
            if os.path.exists(fname+"x"):
                #srcf = fname
                shutil.move(fname,dst)
    else:
        print "in the conv function.."
        for e in xfl:
            os.chdir(e[:e.rfind('/')].replace('/','//'))
            fname = os.path.join(os.getcwd(),e[e.rfind('/')+1:])
            wb = xl.Workbooks.Open(fname)
            wb.SaveAs(fname+"x", FileFormat = 51)
            wb.Close()
            
    xl.Quit()

def move(src,dst):
    pass
 
def csv_comp(f1,f2):
    w1n = f1[f1.rfind('/')+1:f1.rfind('.')]
    w2n = f2[f2.rfind('/')+1:f2.rfind('.')]
    pass


def the_mess1(l):
    global col1,col2,col3,val1,val2,eps
    col1,col2,col3,val1,val2,eps = l[3],l[4],l[5],l[6],l[7],l[8]
    #print col1,col2,col3,val1,val2,eps
    #print l
    #print "Comparision started...!"
    report = []
    if l[1] != '' and l[2] != '':
        f1 = l[1]
        f2 = l[2]
        #print f1,f2
        if f1.endswith('.xls') and f2.endswith('.xls'):
            xls_x_conv([f1,f2])
            rprt = the_mayhem1(f1,f2)
            report.append(rprt)
        else:
            rprt = the_mayhem1(f1,f2)
            #print rprt
            report.append(rprt)
            #print report
    elif l[0] != '':
        work_dir = l[0].replace("/","//")
        os.chdir(work_dir)
        fold_list = [x.lower() for x in next(os.walk('.'))[1]]
        fold_list.sort()
        if len(fold_list) > 1:
            for e in fold_list:
                os.chdir(work_dir)
                print "now in...directory... " + str(e)
                pw_dir = os.path.join(os.getcwd(),e)
                os.chdir(pw_dir)
                #print pw_dir
                #print os.getcwd()
                tocon_list = filter(lambda x: x.endswith('.xls'), os.listdir(pw_dir))
                if tocon_list:
                    xls_x_conv(tocon_list,pw_dir)
                files_list = [e.lower() for e in filter(lambda x: x.endswith('.xlsx'), os.listdir(pw_dir))]
                files_list.sort()
                lfile = len(files_list)
                #print files_list
                #print os.getcwd()
                #os.chdir(work_dir)
                cnt = 0
                while cnt < lfile:
                    f1 = files_list[cnt]
                    f2 = files_list[cnt+1]
                    cnt += 2
                    rprt = the_mayhem1(f1,f2)
                    report.append(rprt)
        #return report
            
        else:
            tocon_list = filter(lambda x: x.endswith('.xls'), os.listdir(work_dir))
            #print tocon_list
            if tocon_list:
                xls_x_conv(tocon_list,work_dir)
            #print "successfully changed path!"
            files_list = [e.lower() for e in filter(lambda x: x.endswith('.xlsx'), os.listdir(work_dir))]
            files_list.sort()
            lfile = len(files_list)
            #print files_list
            #print os.getcwd()
            cnt = 0
            while cnt < lfile:
                f1 = files_list[cnt]
                f2 = files_list[cnt+1]
                cnt += 2
                rprt = the_mayhem1(f1,f2)
                report.append(rprt)
    return report
        #print report

def the_mayhem1(f1,f2):
    print "Comparing... " +f1 +" vs " + f2
    t1 = time.time()
    w1 = load_workbook(f1)
    w2 = load_workbook(f2)
    w1n = f1[f1.rfind('/')+1:f1.rfind('.')]
    w2n = f2[f2.rfind('/')+1:f2.rfind('.')]
    wbsucc = "Pass"
    sdic = {}
    print col1,col2,col3,val1,val2,eps
    
    for i in xrange(len(w1.worksheets)):
        for j in xrange(i,i+1):
            smsg = compare(w1.worksheets[j], w2.worksheets[j],w1n,w2n)
            sdic[w1.worksheets[j].title] = smsg[0],smsg[1]
            w2.save(f2)
    for e in sdic.values():
        if "Pass" not in e[0]:
            wbsucc = "Fail"
    
    t2 = time.time()
    ttime = round((t2-t1),2)
    #print sdic
    return [ttime,w1n,w2n,wbsucc,sdic]
   


if __name__ == '__main__':
    main()

